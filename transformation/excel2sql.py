#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/3/17 15:11
# @Author  : Wendyltanpcy
# @File    : excel2sql.py
# @Software: PyCharm

'''
This code uses the openpyxl package for playing around with excel using Python code
to convert complete excel workbook (all sheets) to an SQLite database
The code assumes that the first row of every sheet is the column name
Every sheet is stored in a separate table
The sheet name is assigned as the table name for every sheet
'''

import sqlite3
from openpyxl import load_workbook
import re
from openpyxl.compat import unicode

def slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text


def doTransfer(file_name,db_name):
    # Replace with a database name
    con = sqlite3.connect(db_name)
    # replace with the complete path to youe excel workbook
    wb = load_workbook(filename=file_name)

    sheets = wb.sheetnames

    for sheet in sheets:
        ws = wb[sheet]

        columns = []
        query = 'CREATE TABLE ' + str(slugify(sheet)) + '(ID INTEGER PRIMARY KEY AUTOINCREMENT'
        for row in next(ws.rows):
            query += ', ' + slugify(row.value) + ' TEXT'
            columns.append(slugify(row.value))
        query += ');'

        con.execute(query)

        tup = []
        for i, rows in enumerate(ws):
            tuprow = []
            if i == 0:
                continue
            for row in rows:
                tuprow.append(unicode(row.value).strip()) if unicode(row.value).strip() != 'None' else tuprow.append('')
            tup.append(tuple(tuprow))

        insQuery1 = 'INSERT INTO ' + str(slugify(sheet)) + '('
        insQuery2 = ''
        for col in columns:
            insQuery1 += col + ', '
            insQuery2 += '?, '
        insQuery1 = insQuery1[:-2] + ') VALUES('
        insQuery2 = insQuery2[:-2] + ')'
        insQuery = insQuery1 + insQuery2

        con.executemany(insQuery, tup)
        con.commit()

    con.close()

